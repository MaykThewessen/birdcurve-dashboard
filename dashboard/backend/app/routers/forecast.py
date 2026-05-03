"""DA and ID3/Imbalance forecasts, annual statistics."""
from __future__ import annotations

import os
from datetime import datetime, timezone
from functools import lru_cache

import openpyxl
import pandas as pd
from fastapi import APIRouter, Query, Request, HTTPException, Response
from fastapi.concurrency import run_in_threadpool

from ..downsampling import lttb_downsample
from ._helpers import get_engine_and_dir, add_cache_headers, to_utc_iso

router = APIRouter(prefix="/forecast", tags=["forecast"])


def _today_iso() -> str:
    return datetime.now(timezone.utc).date().isoformat()


# Resampling rules per resolution. None = take rows as-is at native granularity.
_RESAMPLE_RULES: dict[str, str | None] = {
    "15min": None,
    "hourly": "1h",
    "daily": "1D",
}


def _auto_resolution(start: str, end: str) -> str:
    try:
        d_start = datetime.fromisoformat(start[:10])
        d_end = datetime.fromisoformat(end[:10])
        days = (d_end - d_start).days
    except ValueError:
        return "hourly"
    if days <= 7:
        return "15min"
    if days <= 90:
        return "hourly"
    return "daily"


@lru_cache(maxsize=4)
def _read_da_forecast_csv(csv_path: str, mtime: float) -> pd.DataFrame:
    """Load the predictions_DA_hourly CSV once per (path, mtime).

    The 'mtime' arg forces invalidation when the file on disk changes —
    same trick used by lru_cache_with_invalidation patterns.
    """
    # Handle both 'datetime_UTC' (v17+) and 'datetime' (older) column names.
    df = pd.read_csv(csv_path)
    dt_col = "datetime_UTC" if "datetime_UTC" in df.columns else "datetime"
    df[dt_col] = pd.to_datetime(df[dt_col]).dt.tz_localize("UTC")
    df = df.rename(columns={dt_col: "datetime_UTC"})
    return df


def _get_da_forecast_sync(fdir, start: str, end: str, max_points: int, resolution: str):
    csv_paths = list(fdir.glob("predictions_DA_hourly_*.csv"))
    if not csv_paths:
        raise HTTPException(404, "predictions_DA_hourly CSV not found")
    csv_path = str(csv_paths[0])
    df = _read_da_forecast_csv(csv_path, os.path.getmtime(csv_path))

    # Filter by range. Range strings are dates ('YYYY-MM-DD'); compare against
    # the tz-aware datetime column by promoting them to UTC tz-aware Timestamps.
    start_ts = pd.Timestamp(start, tz="UTC")
    end_ts = pd.Timestamp(end, tz="UTC")
    mask = (df["datetime_UTC"] >= start_ts) & (df["datetime_UTC"] <= end_ts)
    df = df.loc[mask, ["datetime_UTC", "Price_actual", "Price_pred_ensemble"]]

    rule = _RESAMPLE_RULES.get(resolution)
    if rule is not None and not df.empty:
        df = (
            df.set_index("datetime_UTC")
              .resample(rule)
              .mean(numeric_only=True)
              .dropna(how="all")
              .reset_index()
        )

    data = [
        {
            "datetime": to_utc_iso(str(r.datetime_UTC)),
            "price_actual": None if pd.isna(r.Price_actual) else float(r.Price_actual),
            "price_predicted": None if pd.isna(r.Price_pred_ensemble) else float(r.Price_pred_ensemble),
        }
        for r in df.itertuples(index=False)
    ]

    if len(data) > max_points:
        # LTTB needs a numeric y-axis; fall back to price_actual if predicted is None.
        for i, d in enumerate(data):
            d["_idx"] = i
            d["_y"] = d["price_predicted"] if d["price_predicted"] is not None else (d["price_actual"] or 0.0)
        data = lttb_downsample(data, "_idx", "_y", max_points)
        for d in data:
            d.pop("_idx", None)
            d.pop("_y", None)

    return {
        "datetime": [d["datetime"] for d in data],
        "price_actual": [d["price_actual"] for d in data],
        "price_predicted": [d["price_predicted"] for d in data],
    }


@router.get("/da")
async def get_da_forecast(
    request: Request,
    response: Response,
    start: str = Query(...),
    end: str = Query(...),
    scenario: str = Query(...),
    max_points: int = Query(10000, ge=10, le=100000),
    resolution: str = Query("auto", pattern="^(auto|15min|hourly|daily)$"),
):
    _engine, fdir = get_engine_and_dir(request, scenario)
    if resolution == "auto":
        resolution = _auto_resolution(start, end)
    payload = await run_in_threadpool(
        _get_da_forecast_sync, fdir, start, end, max_points, resolution
    )
    add_cache_headers(response, end, _today_iso())
    return payload


def _get_id3_imbalance_sync(engine, scenario: str, start: str, end: str, max_points: int):
    data = engine.query_forecast_file(
        scenario,
        "predictions_DA_ID3_Imb_aFRR_FCR_quarterly_2023_2050",
        start, end,
        datetime_col="Datetime_UTC",
    )

    def _safe(v):
        """Convert NaN/NA pandas values to None for JSON serialization."""
        import math
        if v is None:
            return None
        try:
            if math.isnan(float(v)):
                return None
        except (TypeError, ValueError):
            pass
        return v

    result = []
    for d in data:
        result.append({
            "datetime": str(d.get("Datetime_UTC", "")),
            "da_price": _safe(d.get("Day-ahead_price")),
            "id3_price": _safe(d.get("ID3_price")),
            "afrr_up": _safe(d.get("aFRR_Energy_up_price")),
            "afrr_down": _safe(d.get("aFRR_Energy_down_price")),
            "imb_long": _safe(d.get("imb_long_price")),
            "imb_short": _safe(d.get("imb_short_price")),
            "reg_state": _safe(d.get("Regulation_State")),
        })

    if len(result) > max_points:
        for i, d in enumerate(result):
            d["_idx"] = i
        result = lttb_downsample(result, "_idx", "da_price", max_points)
        for d in result:
            d.pop("_idx", None)

    return {
        "datetime": [r["datetime"] for r in result],
        "da_price": [r["da_price"] for r in result],
        "id3_price": [r["id3_price"] for r in result],
        "afrr_up": [r["afrr_up"] for r in result],
        "afrr_down": [r["afrr_down"] for r in result],
        "imb_long": [r["imb_long"] for r in result],
        "imb_short": [r["imb_short"] for r in result],
        "reg_state": [r["reg_state"] for r in result],
    }


@router.get("/id3-imbalance")
async def get_id3_imbalance(
    request: Request,
    response: Response,
    start: str = Query(...),
    end: str = Query(...),
    scenario: str = Query(...),
    max_points: int = Query(10000, ge=10, le=100000),
):
    engine, _fdir = get_engine_and_dir(request, scenario)
    payload = await run_in_threadpool(
        _get_id3_imbalance_sync, engine, scenario, start, end, max_points
    )
    add_cache_headers(response, end, _today_iso())
    return payload


def _get_annual_stats_sync(fdir):
    xlsx_files = list(fdir.glob("Annual_statistics_*.xlsx"))
    if not xlsx_files:
        raise HTTPException(404, "Annual_statistics xlsx not found")

    wb = openpyxl.load_workbook(xlsx_files[0], read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_data = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows_data:
        return {}

    years = [int(y) for y in rows_data[0][1:] if y is not None]

    metric_map = {}
    for row in rows_data[1:]:
        if row[0] is None:
            continue
        metric_name = str(row[0])
        values = [float(v) if v is not None else 0.0 for v in row[1:len(years) + 1]]
        metric_map[metric_name] = values

    return {
        "years": years,
        "avg_da": metric_map.get("Avg DA Price (€/MWh)", []),
        "std_da": metric_map.get("Std Dev DA Price (€/MWh)", []),
        "spread": metric_map.get("Avg Daily Spread (€/MWh)", []),
        "negative_hours": metric_map.get("Negative Hours (<€0)", []),
        "peak_hours": metric_map.get("Peak Hours (>€200)", []),
        "bess_2h": metric_map.get("BESS 2h DA Revenue (k€/MW/y)", []),
        "bess_4h": metric_map.get("BESS 4h DA Revenue (k€/MW/y)", []),
        "bess_8h": metric_map.get("BESS 8h DA Revenue (k€/MW/y)", []),
        "bess_id3": metric_map.get("BESS 2h ID3 Revenue (k€/MW/y)", []),
        "bess_afrr": metric_map.get("BESS 2h aFRR Energy Revenue (k€/MW/y)", []),
        "afrr_cap_rev": metric_map.get("aFRR Capacity Revenue (k€/MW/y)", []),
        "fcr_cap_rev": metric_map.get("FCR Capacity Revenue (k€/MW/y)", []),
        "solar_capture": metric_map.get("Solar Capture Price (€/MWh)", []),
        "solar_rev": metric_map.get("Solar PV Revenue (k€/MW/y)", []),
        "wind_rev": metric_map.get("Wind Revenue (k€/MW/y)", []),
        "demand_twh": metric_map.get("Annual Demand (TWh/y)", []),
    }


@router.get("/annual-stats")
async def get_annual_stats(
    request: Request,
    response: Response,
    scenario: str = Query(...),
):
    _engine, fdir = get_engine_and_dir(request, scenario)
    payload = await run_in_threadpool(_get_annual_stats_sync, fdir)
    # Annual stats span all years incl. future projections → recent.
    response.headers["Cache-Control"] = "public, max-age=300"
    return payload
